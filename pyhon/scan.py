import openpyxl 
import datetime
from openpyxl import Workbook

x =datetime.datetime.now()
autows = x.strftime("%B")
wb = openpyxl.load_workbook('C:\\Users\\SCDSN0\\PycharmProjects\\SCDSNproject\\Excle\\คะแนนความสะอาด\\Term1\\Term1.xlsx')
ws = wb.active

def scanrow():
	countrow = 0
	for row in ws.iter_rows(min_row=1,max_row=15,min_col=1,max_col=1):
		for cell in row:
			countrow += 1
			if cell.value == '0110':
				return countrow

def scancol():
	countcol = 96
	time = datetime.datetime.now()
	timecal = str(time.year) + '-' + str(time.month) + '-' + str(time.day)
	
	for row in ws.iter_rows(min_row=1,max_row=1,min_col=1,max_col=15):
		for cell in row:
			x = str(cell.value)
			y = x.replace(' 00:00:00','')
			countcol += 1
			if y == timecal:
				return chr(countcol)

def scanroom(room):				
	row = scanrow(room)
	col = scancol()
	cell = str(col)+chr(row)
	print('cell is ' + cell)
	return cell


row = scanrow()
col = scancol()
cell = str(col)+str(row)
print('cell is ' + cell)
x = int(input('Enter score: '))
ws[str(cell)] = x

wb.save('C:\\Users\\SCDSN0\\PycharmProjects\\SCDSNproject\\Excle\\คะแนนความสะอาด\\Term1\\Term1.xlsx')